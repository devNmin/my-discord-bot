# discord 라이브러리 사용 선언
import discord
import random
import openpyxl

class chatbot(discord.Client):
    # 프로그램이 처음 실행되었을 때 초기 구성
    async def on_ready(self):
        # 상태 메시지 설정
        # 종류는 3가지: Game, Streaming, CustomActivity
        bot_message = "봇 " # ~~하는중
        game = discord.Game(f"{bot_message}")
        
        # 계정 상태를 변경한다=> 온라인 상태, game 중으로 설정
        await client.change_presence(status=discord.Status.online, activity=game)

        # 준비가 완료되면 콘솔 창에 "READY!"라고 표시
        print("준비완료")


    # 봇에 메시지가 오면 수행 될 액션
    async def on_message(self, message):

        # SENDER가 BOT일 경우 반응을 하지 않도록 한다.
        if message.author.bot:
            return None

        # 현재 채널 불러오기
        channel = message.channel
        # 기억하기
        if message.content.startswith("$학습") or message.content.startswith("$기억"):
            # memory 엑셀파일 불러오기
            file = openpyxl.load_workbook("memory.xlsx")
            sheet = file['Sheet1']
            
            learn  = message.content.split(" ",maxsplit=2)
            
            for i in range(1,50): 

                if  sheet[f"A{i}"].value == learn[1]: # 값이 있으면 덮어쓰기
                    sheet[f"A{i}"].value, sheet[f"B{i}"].value = learn[1], learn[2]
                    await channel.send(f'```{learn[1]} 이것을 덮어쓰기.```')
                    break

                elif sheet[f"A{i}"].value == None : #값이 없으면 새로쓰기
                    sheet[f"A{i}"].value, sheet[f"B{i}"].value  = learn[1], learn[2]
                    await channel.send(f'```{learn[1]} 이것을 기억.```')
                    break
               
            file.save("memory.xlsx")
            return

        # 알려주기        
        elif message.content.startswith('$알려'):
            file = openpyxl.load_workbook('memory.xlsx') # sheet 읽어오기 
            sheet = file['Sheet1']
            memory  = message.content.split(" ")
            
            for i in range(1,50):

                if sheet[f"A{i}"].value == memory[1]: 
                    await channel.send(f'```{sheet["B"+str(i)].value}```')
                    break
        # 삭제       
        elif message.content.startswith('$삭제'):
            # sheet 읽어오기 
            file = openpyxl.load_workbook('memory.xlsx')
            sheet = file['Sheet1']
            memory  = message.content.split(" ")
            
            for i in range(1,50):
                if sheet[f"A{i}"].value == memory[1]:
                    sheet.delete_rows(i)

                    await channel.send(f'```{memory[1]}가 삭제되었습니다.```')
                    break
            file.save("memory.xlsx")
            return

        

        if message.content == "반가워요":
            # 답변 내용 구성
            msg = "저도 반갑습니다."
            # msg에 지정된 내용대로 메시지를 전송
            await channel.send(msg)
            return None



        elif message.content == "주사위":

            msg = random.randint(1,6)
            # msg에 지정된 내용대로 메시지를 전송
            await channel.send(msg)
            return None

        elif message.content == "로또":
            rotto = list(range(1,46))
            random.shuffle(rotto)
            msg = rotto[0:6]
            # msg에 지정된 내용대로 메시지를 전송
            await channel.send(msg)
            return None





# 프로그램이 실행되면 제일 처음으로 실행되는 함수
if __name__ == "__main__":
    # 객체를 생성
    client = chatbot()

    # TOKEN 값을 통해 로그인하고 봇을 실행
    Token_number = ""
    client.run(Token_number)